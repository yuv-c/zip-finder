import asyncio
from openpyxl import load_workbook, cell, worksheet, workbook
from typing import Tuple, Iterable
import logging
import time
from es_client import ElasticSearchClient  # import the ElasticSearch client

LOG_FORMAT = "%(asctime)s | %(levelname)s | %(name)s | %(lineno)d | %(message)s"
logging.basicConfig(level=logging.INFO, format=LOG_FORMAT)

ZIP_CODES_FILE = "zip-codes-small.xlsx"
COL_INDEX_TO_NAME = {
    1: "LocationID",
    2: "City Name",
    3: "StreetID",
    4: "Street Name",
    5: "House Number",
    6: "Entrance",
    7: "ZIP 7",
    8: "Remark",
    9: "Updated",
}

"""
interesting row props:
every row is a tuple of cells

every cell has:
col_idx: integer
row: integer
value: probably str
"""

INDEX = "address-to-zip"


def get_excel_file_handles(file_path: str) -> (worksheet, workbook):
    work_book = load_workbook(filename=file_path, read_only=True)
    return work_book.active, work_book


class BadDataError(Exception):
    pass


def verify_street_or_city_is_valid(street_or_city: str) -> None:
    if not street_or_city:
        raise BadDataError
    if street_or_city == "?":
        raise BadDataError
    try:
        str(street_or_city)
    except ValueError:
        raise BadDataError


def verify_numbers(number: str) -> None:
    if not number or len(number) < 5:
        raise BadDataError

    try:
        int(number)
    except ValueError:
        raise BadDataError


def yield_rows_from_work_sheet(
        work_sheet: worksheet, min_row: int, max_row: int
) -> Iterable[
    Tuple[
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
    ]
]:
    for row in work_sheet.iter_rows(min_row=min_row, max_row=max_row):
        if row[0].value is None:
            break
        yield row


async def log_bad_address_to_file(
        row: Tuple[
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
        ]
) -> None:
    # You didn't think I'd actually implement this, did you?
    pass


async def print_cell_content(
        row_data: Tuple[
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
        ]
) -> None:
    (
        loc_id,
        city_name,
        street_id,
        street_name,
        house_num,
        entrance,
        zip_code,
        remark,
        updated,
    ) = row_data
    row_num = row_data[0].row

    try:
        list(map(verify_street_or_city_is_valid, (city_name.value, street_name.value)))
        list(
            map(
                verify_numbers,
                (house_num.value, street_id.value, loc_id.value, zip_code.value),
            )
        )
    except BadDataError:
        logging.warning(f"Row {row_num}: Invalid data, skipping")
        return

    print(
        f"Row {row_num}: {loc_id.value},"
        f" {city_name.value}, {street_id.value}, {street_name.value}, {house_num.value}, "
        f"{entrance.value}, {zip_code.value}, {remark.value}, {updated.value}"
    )


def cell_content_to_dict(
        row_data: Tuple[
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
            cell.Cell,
        ]
) -> dict:
    (
        loc_id,
        city_name,
        street_id,
        street_name,
        house_num,
        entrance,
        zip_code,
        remark,
        updated,
    ) = row_data
    row_num = row_data[0].row

    try:
        list(map(verify_street_or_city_is_valid, (city_name.value, street_name.value)))
        list(
            map(
                verify_numbers,
                (house_num.value, street_id.value, loc_id.value, zip_code.value),
            )
        )
    except BadDataError:
        logging.warning(f"Row {row_num}: Invalid data, skipping")
        return {}

    return {
        "row_num": row_num,
        "loc_id": loc_id.value,
        "city_name": city_name.value,
        "street_id": street_id.value,
        "street_name": street_name.value,
        "house_number": house_num.value,
        "entrance": entrance.value,
        "zip_code": zip_code.value,
        "remark": remark.value,
        "updated": updated.value
    }


async def create_index(es_client: ElasticSearchClient) -> None:
    mapping = {
        "mappings": {
            "properties": {
                "row_num": {"type": "integer"},
                "loc_id": {"type": "integer"},
                "city_name": {"type": "text",
                              "analyzer": "hebrew"},
                "street_id": {"type": "integer"},
                "street_name": {"type": "text",
                                "analyzer": "hebrew"},
                "house_number": {"type": "text"},
                "entrance": {"type": "text",
                             "analyzer": "hebrew"},
                "zip_code": {"type": "integer"},
                "remark": {"type": "text",
                           "analyzer": "hebrew"},
                "updated": {"type": "date",
                            "format": "yyyyMMdd"}
            }
        }
    }
    es_client.create_index(index=INDEX, body=mapping)


async def main() -> None:
    es_client = ElasticSearchClient(es_endpoint="localhost", es_index=INDEX,
                                    es_port=9200)  # use the ElasticSearchClient

    await create_index(es_client=es_client)

    rows_step_size = 1000  # Window size
    work_sheet, open_workbook = get_excel_file_handles(file_path=ZIP_CODES_FILE)
    rows_number = work_sheet.max_row
    logging.info(f"Rows number: {rows_number}")

    min_row = 2  # skip header
    window_min_row = min_row
    window_max_row = window_min_row + rows_step_size - 1
    for _ in range(min_row, rows_number, rows_step_size):
        logging.info(f"Processing window: {window_min_row} - {window_max_row}")
        list_to_push = []
        for row in yield_rows_from_work_sheet(
                work_sheet=work_sheet, min_row=window_min_row, max_row=window_max_row
        ):

            if row[0].value is not None:  # EOF
                try:
                    list_to_push.append(cell_content_to_dict(row_data=row))
                except BadDataError:
                    await log_bad_address_to_file(row=row)

            else:
                logging.info("EOF")
                break

        window_min_row += rows_step_size
        window_max_row += rows_step_size

        es_client.bulk_push_to_elasticsearch(list_of_docs=list_to_push, index=INDEX)
        logging.info(f"Finished processing window: {window_min_row} - {window_max_row}")

    open_workbook.close()


if __name__ == "__main__":
    start = time.perf_counter()
    asyncio.run(main=main())
    print(f"{__file__} executed in {time.perf_counter() - start:0.2f} seconds.")