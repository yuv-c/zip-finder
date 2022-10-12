import asyncio
from openpyxl import load_workbook, cell, worksheet, workbook
from typing import Tuple, Iterable
import logging
import time
from opensearchpy import OpenSearch


LOG_FORMAT = "%(asctime)s | %(levelname)s | %(name)s | %(lineno)d | %(message)s"
logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT)

ZIP_CODES_FILE = "zip-codes.xlsx"
COL_INDEX_TO_NAME = {
    1: "LocationID",
    2: "City Name",
    3: "StreetID",
    4: "Street Name",
    5: "House Number",
    6: "Entrance",
    7: "ZIP 7",
    8: "Remark",
    9: "Updated",
}

"""
interesting row props:
every row is a tuple of cells

every cell has:
col_idx: integer
row: integer
value: probably str
"""

AWS_DEFAULT_REGION = "eu-central-1"

def get_excel_file_handles(file_path: str) -> (worksheet, workbook):
    work_book = load_workbook(filename=file_path, read_only=True)
    return work_book.active, work_book


class BadDataError(Exception):
    pass


def verify_street_or_city_is_valid(street_or_city: str) -> None:
    if not street_or_city:
        raise BadDataError
    if street_or_city == "?":
        raise BadDataError
    try:
        str(street_or_city)
    except ValueError:
        raise BadDataError


def verify_numbers(number: str) -> None:
    if not number or len(number) < 5:
        raise BadDataError

    try:
        int(number)
    except ValueError:
        raise BadDataError


def yield_rows_from_work_sheet(
    work_sheet: worksheet, min_row: int, max_row: int
) -> Iterable[
    Tuple[
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
    ]
]:
    for row in work_sheet.iter_rows(min_row=min_row, max_row=max_row):
        if row[0].value is None:
            break
        yield row


async def log_bad_address_to_file(
    row: Tuple[
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
    ]
) -> None:
    # You didn't think I'd actually implement this, did you?
    pass


async def print_cell_content(
    row_data: Tuple[
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
        cell.Cell,
    ]
) -> None:
    (
        loc_id,
        city_name,
        street_id,
        street_name,
        house_num,
        entrance,
        zip_code,
        remark,
        updated,
    ) = row_data
    row_num = row_data[0].row

    try:
        list(map(verify_street_or_city_is_valid, (city_name.value, street_name.value)))
        list(
            map(
                verify_numbers,
                (house_num.value, street_id.value, loc_id.value, zip_code.value),
            )
        )
    except BadDataError:
        logging.warning(f"Row {row_num}: Invalid data, skipping")
        return

    print(
        f"Row {row_num}: {loc_id.value},"
        f" {city_name.value}, {street_id.value}, {street_name.value}, {house_num.value}, "
        f"{entrance.value}, {zip_code.value}, {remark.value}, {updated.value}"
    )


async def main() -> None:
    rows_step_size = 1000
    work_sheet, open_workbook = get_excel_file_handles(file_path=ZIP_CODES_FILE)
    rows_number = work_sheet.max_row
    logging.info(f"Rows number: {rows_number}")

    rows_number = 10000  # TODO: Remove

    min_row = 1
    max_row = 1000

    for number in range(min_row, rows_number, rows_step_size):
        for row in yield_rows_from_work_sheet(
            work_sheet=work_sheet, min_row=min_row, max_row=max_row
        ):

            if row[0].value is None:  # EOF
                break

            try:
                await print_cell_content(row_data=row)
            except BadDataError:
                await log_bad_address_to_file(row=row)
                continue

        min_row += rows_step_size
        max_row += rows_step_size

    open_workbook.close()


if __name__ == "__main__":
    start = time.perf_counter()
    asyncio.run(main=main())
    print(f"{__file__} executed in {time.perf_counter() - start:0.2f} seconds.")

